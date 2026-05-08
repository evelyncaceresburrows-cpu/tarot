# -*- coding: utf-8 -*-
"""Vuelca el Excel de Arcanos Mayores a un JSON consultable
   en src/content/sourceMaterial.json. Esta es la fuente
   estructural que ADE puede consultar para reescribir cartas."""
import openpyxl
import json
import os

PATH = r'E:\CARTAS\docs\rider-waite.xlsx'
OUT  = r'E:\CARTAS\src\content\sourceMaterial.json'

wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb['Table 1']

headers = [str(c.value or '').strip() for c in ws[1]]
KEYMAP = {
    'Número del Arcano':        'number',
    'Nombre de la Carta':       'name',
    'Significado Simbólico':    'symbolicMeaning',
    'Atribución de la Golden Dawn': 'goldenDawn',
    'Simbolismo de la Imagen':  'imageSymbolism',
    'Concepto de Waite':        'waite',
    'Referencia Histórica o Bíblica': 'historicalReference',
    'Fuente':                   'sourceRefs'
}

cards = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row): continue
    obj = {}
    for h, v in zip(headers, row):
        key = KEYMAP.get(h, h)
        obj[key] = (str(v).strip() if v is not None else '')
    cards.append(obj)

# Source references
sourcesWS = wb['Source References']
sources = {}
for row in sourcesWS.iter_rows(min_row=2, values_only=True):
    if row[0] is not None:
        sources[str(row[0])] = row[1]

out = {
    'cards': cards,
    'sourceReferences': sources,
    'note': 'Fuente: Excel "Simbolismo y Atributos de los Arcanos Mayores Rider-Waite". Material de consulta interno para reescribir interpretaciones con anclaje visual y simbólico concreto.'
}

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

print(f'OK · {len(cards)} cartas en {OUT}')
print(f'   ({len(sources)} referencias bibliográficas)')
