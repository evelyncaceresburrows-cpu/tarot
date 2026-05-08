# -*- coding: utf-8 -*-
"""Vuelca la Guía Analítica a JSON consultable."""
import openpyxl
import json
import os
import re

PATH = r'E:\CARTAS\docs\guia-analitica-arcanos.xlsx'
OUT  = r'E:\CARTAS\src\content\analyticGuide.json'

wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb['Table 1']

headers = [str(c.value or '').strip() for c in ws[1]]
KEYMAP = {
    'Carta':                                            'name',
    'Número':                                           'number',
    'Imagen simbólica principal':                       'imageMain',
    'Elementos visuales importantes':                   'visualElements',
    'Contradicción humana central':                     'centralContradiction',
    'Tensión emocional principal':                      'emotionalTension',
    'Movimiento interno de la carta':                   'innerMovement',
    'Situaciones humanas reconocibles (3)':             'humanSituationsRaw',
    'Sombra concreta de la carta':                      'shadow',
    'Tipo de conflicto humano':                         'conflictType',
    'Sensación atmosférica':                            'atmosphere',
    'Ritmo emocional':                                  'rhythm',
    'Dirección emocional':                              'direction',
    'Símbolos secundarios importantes':                 'secondarySymbols',
    'Riesgo de interpretación cliché':                  'clicheRisk',
    'Qué NO debería decir una interpretación superficial': 'whatNotToSay',
    'Cómo aterrizar la carta emocionalmente':           'howToGround',
    'Diferencia con cartas similares':                  'differentiation',
    'Pregunta reflexiva potente':                       'reflectivePrompt',
    'Interpretación Rider-Waite tradicional resumida':  'riderWaiteSummary'
}

def clean_text(t):
    if t is None: return ''
    s = str(t).strip()
    # quitar marcadores tipo "1, 2." al final que vienen del paper
    s = re.sub(r'\s*\d+(\s*,\s*\d+)*\s*\.?$', '', s).strip()
    return s

def parse_situations(raw):
    """Toma '...trabajo sin tener plan B. 2) Ignorar... 3) Sentir...' y devuelve 3 situaciones limpias."""
    if not raw: return []
    s = str(raw).strip()
    # quitar marcadores de fuente
    s = re.sub(r'\s*\d+(\s*,\s*\d+)*\s*\.?$', '', s).strip()
    # split por '2)', '3)' o por '1)' inicial
    # primero dividir por puntos seguidos de espacio y número)
    parts = re.split(r'\s*\d+\)\s*', s)
    # algunas situaciones están separadas con punto y luego "2)"
    # también puede empezar sin "1)"
    cleaned = [p.strip().rstrip('.').strip() for p in parts if p.strip()]
    # Asegurar 3
    return cleaned[:3]


cards = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row): continue
    obj = {}
    for h, v in zip(headers, row):
        key = KEYMAP.get(h, h)
        obj[key] = clean_text(v)
    obj['humanSituations'] = parse_situations(obj.get('humanSituationsRaw', ''))
    cards.append(obj)

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump({'cards': cards}, f, ensure_ascii=False, indent=2)

print(f'OK · {len(cards)} cartas en {OUT}')
print('\nMUESTRA situaciones parsed:')
for c in cards[:3]:
    print(f'\n{c["name"]}:')
    for i, s in enumerate(c['humanSituations']):
        print(f'  {i+1}. {s}')
