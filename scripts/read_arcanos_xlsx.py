# -*- coding: utf-8 -*-
"""Lee el Excel de Simbolismo Rider-Waite y vuelca su contenido."""
import openpyxl
import sys

PATH = r'E:\CARTAS\docs\rider-waite.xlsx'
wb = openpyxl.load_workbook(PATH, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n{"="*70}')
    print(f'HOJA: {sheet_name}  ({ws.max_row} filas × {ws.max_column} cols)')
    print('='*70)
    headers = [str(c.value or '').strip() for c in ws[1]]
    print('\nCOLUMNAS:')
    for i, h in enumerate(headers):
        print(f'  [{i}] {h}')

    # Primera fila completa
    if ws.max_row >= 2:
        print('\n--- PRIMERA FILA DE DATOS ---')
        for h, v in zip(headers, [c.value for c in ws[2]]):
            if v:
                print(f'\n[{h}]')
                print(f'  {str(v)[:600]}')

    # Cuántas filas tienen datos
    print(f'\n--- {ws.max_row - 1} filas de datos en total ---')
